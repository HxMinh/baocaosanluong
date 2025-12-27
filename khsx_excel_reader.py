# -*- coding: utf-8 -*-
"""
Excel Reader Module - Read password-protected Excel files
"""

import io
import os
import datetime
import msoffcrypto
import openpyxl
import pandas as pd
from typing import Dict, List, Optional


def read_excel_with_password(
    file_path: str, 
    password: str,
    target_sheets: Optional[List[str]] = None
) -> Optional[Dict[str, pd.DataFrame]]:
    """
    Read password-protected Excel file and return specified sheets as DataFrames
    
    Args:
        file_path: Path to Excel file
        password: Password to decrypt file
        target_sheets: List of sheet names to read (None = read all sheets)
        
    Returns:
        Dictionary of {sheet_name: DataFrame} or None if error
    """
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            print("[ERROR] File not found")
            return None
        
        # Open encrypted file
        print("[INFO] Opening encrypted file...")
        with open(file_path, 'rb') as f:
            file = msoffcrypto.OfficeFile(f)
            file.load_key(password=password)
            
            # Decrypt to memory
            print("[INFO] Decrypting file (this may take a while for large files)...")
            decrypted = io.BytesIO()
            file.decrypt(decrypted)
            
            # Read Excel from memory
            print("[INFO] Loading workbook...")
            decrypted.seek(0)
            workbook = openpyxl.load_workbook(decrypted, data_only=True)
            
            # Convert specified sheets to DataFrames
            sheets_data = {}
            
            # Determine which sheets to read
            sheets_to_read = target_sheets if target_sheets else workbook.sheetnames
            print(f"[INFO] Found {len(workbook.sheetnames)} sheets total, reading {len(sheets_to_read)} target sheet(s)...")
            
            for idx, sheet_name in enumerate(sheets_to_read, 1):
                # Skip if sheet doesn't exist
                if sheet_name not in workbook.sheetnames:
                    print(f"[WARNING] Sheet '{sheet_name}' not found in workbook, skipping...")
                    continue
                
                print(f"[INFO] Reading sheet {idx}/{len(sheets_to_read)}: '{sheet_name}'...")
                sheet = workbook[sheet_name]
                
                # Get all values
                print(f"[INFO]   Processing rows (this may take several minutes for large sheets)...")
                data = []
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    # Convert datetime objects to DD/MM/YYYY format
                    converted_row = []
                    for cell in row:
                        if isinstance(cell, (pd.Timestamp, datetime.datetime, datetime.date)):
                            # Format as DD/MM/YYYY (date only, no time)
                            if isinstance(cell, datetime.datetime):
                                converted_row.append(cell.strftime('%d/%m/%Y'))
                            elif isinstance(cell, datetime.date):
                                converted_row.append(cell.strftime('%d/%m/%Y'))
                            elif isinstance(cell, pd.Timestamp):
                                converted_row.append(cell.strftime('%d/%m/%Y'))
                            else:
                                converted_row.append(str(cell))
                        elif isinstance(cell, datetime.time):
                            # Keep time as is
                            converted_row.append(str(cell))
                        else:
                            converted_row.append(cell)
                    data.append(tuple(converted_row))
                    row_count += 1
                    # Progress feedback every 1000 rows
                    if row_count % 1000 == 0:
                        print(f"[INFO]   Processed {row_count} rows...", end='\r')
                
                if len(data) > 0:
                    # First row as header
                    print(f"\n[INFO]   Converting to DataFrame...")
                    df = pd.DataFrame(data[1:], columns=data[0])
                    sheets_data[sheet_name] = df
                    print(f"[OK] Read sheet '{sheet_name}': {len(df)} rows x {len(df.columns)} columns")
            
            return sheets_data
            
    except PermissionError:
        print(f"[WARNING] File is locked (Excel may be open): {file_path}")
        return None
    except Exception as e:
        print(f"[ERROR] Error reading Excel: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_file_modification_time(file_path: str) -> Optional[float]:
    """
    Get file modification timestamp
    
    Args:
        file_path: Path to file
        
    Returns:
        Modification timestamp or None if error
    """
    try:
        return os.path.getmtime(file_path)
    except:
        return None


if __name__ == "__main__":
    # Test
    from khsx_sync_config import CONFIG
    
    print("Testing Excel reader...")
    sheets = read_excel_with_password(
        CONFIG['excel_file_path'],
        CONFIG['excel_password'],
        CONFIG['target_sheets']
    )
    
    if sheets:
        print(f"\n[OK] Successfully read {len(sheets)} sheets:")
        for sheet_name, df in sheets.items():
            print(f"  - {sheet_name}: {len(df)} rows x {len(df.columns)} columns")
    else:
        print("\n[ERROR] Failed to read Excel file")
